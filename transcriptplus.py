import os
import re
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Função para abrir janela de seleção de pasta
def select_transcriptions_directory():
    root = Tk()
    root.withdraw()  # Esconde a janela principal
    directory = filedialog.askdirectory(title="Selecione a pasta com as transcrições")
    return directory

# Função para capturar palavras-chave do usuário
def get_keywords(prompt):
    keywords = input(prompt)
    return [word.strip() for word in keywords.split(',')]

# Função para obter nome do arquivo Excel
def get_excel_filename():
    return input("Digite o nome do arquivo Excel (com extensão .xlsx): ")

# Função para ler arquivos de transcrições e analisar palavras-chave
def analyze_transcriptions(directory, keywords, conditional_keywords):
    results = []

    for filename in os.listdir(directory):
        if filename.endswith('.txt'):
            file_path = os.path.join(directory, filename)
            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    content = file.readlines()
                    
                    matched_keywords = []
                    excerpts = []
                    
                    for line in content:
                        for word in keywords:
                            if word.lower() in line.lower():
                                matched_keywords.append(word)
                                excerpts.append(line.strip())
                                
                    for line in content:
                        for word in conditional_keywords:
                            if word.lower() in line.lower() and any(k in line.lower() for k in matched_keywords):
                                matched_keywords.append(word)
                                excerpts.append(line.strip())
                                
                    if matched_keywords:
                        video_name = os.path.splitext(filename)[0]
                        results.append({
                            'Nome do Arquivo': video_name,
                            'Palavras Capturadas': ', '.join(set(matched_keywords)),
                            'Trechos Encontrados': '\n'.join(excerpts)
                        })
                        print(f"Palavras capturadas no arquivo {filename}: {matched_keywords}")

            except Exception as e:
                print(f"Erro ao processar o arquivo {filename}: {e}")

    return results

# Função principal
def main():
    # Selecionar pasta de transcrições
    transcriptions_dir = select_transcriptions_directory()

    if not transcriptions_dir:
        print("Nenhuma pasta selecionada. Saindo.")
        return

    # Capturar palavras-chave e palavras condicionais
    keywords = get_keywords("Digite as palavras-chave separadas por vírgulas: ")
    conditional_keywords = get_keywords("Digite as palavras condicionais separadas por vírgulas: ")

    # Capturar nome do arquivo Excel
    excel_filename = get_excel_filename()

    # Analisar as transcrições
    transcription_analysis = analyze_transcriptions(transcriptions_dir, keywords, conditional_keywords)

    if transcription_analysis:
        # Salvar análise em arquivo Excel
        df = pd.DataFrame(transcription_analysis)
        df.to_excel(excel_filename, index=False)
        print(f"Análise concluída. Resultados salvos em {excel_filename}")

        # Aplicar formatação no Excel
        wb = load_workbook(excel_filename)
        ws = wb.active

        # Configurar estilos
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        content_font = Font(size=11)
        alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        border_style = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        # Estilizar cabeçalhos
        for cell in ws["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border_style

        # Estilizar conteúdo das células
        for row in ws.iter_rows(min_row=2, max_col=3):
            for cell in row:
                cell.font = content_font
                cell.alignment = alignment
                cell.border = border_style

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Obter a letra da coluna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(excel_filename)
        print(f"Formatação aplicada. Resultados finais salvos em {excel_filename}")
    else:
        print("Nenhuma palavra-chave foi encontrada nas transcrições.")

if __name__ == "__main__":
    main()